from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from .models import DailyExpenseLine, OtherExpense, TravelIntake


# ---------------------------------------------------------------------------
# Actual 133-103 template column layout (verified from template file):
#
# Header rows 4-7:
#   Row 4 labels: A4=Name, AY4=Employee ID, CH4=Official Station
#   Row 5 data:   A5=name, AY5=emp_id, CH5=official_station
#   Row 6 labels: A6=Address, AY6=City, BP6=State, BU6=Zip, CH6=Official Residence
#   Row 7 data:   A7=address, AY7=city, BP7=state, BU7=zip, CH7=official_residence
#
# Expense rows 10-21 (page 1), 48-80 (page 2):
#   A=date, G=day, J=from, X=to, AL=depart, AR=return,
#   AX=breakfast, AZ=lunch, BB=dinner, BD=lodging,
#   BK=miles, BP=mileage_rate, BU=pov_reason,
#   CB=formula(BK*BP, don't write), CH=other, CN=formula(total, don't write),
#   CU=purpose
#
# Totals / metadata:
#   A28=last_date, Z28=regular_work_hours, A30=remarks
#   BV27=travel_advance (formula CN27=CN25-BV27)
#   AI43=signature_date
#
# Account codes rows 34-38:
#   G=work_order, U=group, AA=org_code, AJ=object_code,
#   BH=work_op, BN=bal_sheet, BT=amount
#
# Other expenses rows 34-39 (right-hand columns):
#   BZ=date, CF=paid_to, CY=for_what, DS=amount
# ---------------------------------------------------------------------------


def _cv(value) -> float:
    return round(float(value or 0), 2)


def _excel_time(value) -> Optional[float]:
    """Convert a time string like '0600' or '06:00' to an Excel time fraction."""
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().upper().replace(".", "")
    normalized = re.sub(r"\s+", " ", text)
    for candidate in [normalized, normalized.replace(" ", "")]:
        for fmt in ("%I:%M%p", "%I%M%p", "%H:%M", "%H%M"):
            try:
                dt = datetime.strptime(candidate, fmt)
                return round((dt.hour * 60 + dt.minute) / 1440, 10)
            except ValueError:
                pass
    return None


def _resolve_cell(ws, cell_ref: str) -> str:
    """Return the top-left cell of any merged range containing cell_ref.

    openpyxl raises AttributeError if you write to a non-top-left merged cell.
    """
    from openpyxl.utils import range_boundaries
    col_letter = cell_ref.rstrip("0123456789")
    row_num = int(cell_ref[len(col_letter):])
    col_num = column_index_from_string(col_letter)
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged))
        if min_col <= col_num <= max_col and min_row <= row_num <= max_row:
            return f"{get_column_letter(min_col)}{min_row}"
    return cell_ref


def _set(ws, cell_ref: str, value) -> None:
    """Write a value to a cell, resolving merged ranges automatically."""
    if value is None or value == "":
        return
    ws[_resolve_cell(ws, cell_ref)] = value


# ---------------------------------------------------------------------------
# Row writer
# ---------------------------------------------------------------------------

def _write_daily_line(ws, row: int, line: DailyExpenseLine) -> None:
    _set(ws, f"A{row}", line.date)
    _set(ws, f"G{row}", line.date.strftime("%a")[:2])
    _set(ws, f"J{row}", line.trip_from)
    _set(ws, f"X{row}", line.trip_to)
    t_dep = _excel_time(line.depart)
    t_ret = _excel_time(line.return_time)
    if t_dep is not None:
        _set(ws, f"AL{row}", t_dep)
    if t_ret is not None:
        _set(ws, f"AR{row}", t_ret)
    if line.breakfast:
        _set(ws, f"AX{row}", _cv(line.breakfast))
    if line.lunch:
        _set(ws, f"AZ{row}", _cv(line.lunch))
    if line.dinner:
        _set(ws, f"BB{row}", _cv(line.dinner))
    if line.lodging:
        _set(ws, f"BD{row}", _cv(line.lodging))
    if line.miles:
        _set(ws, f"BK{row}", _cv(line.miles))
    if line.mileage_rate:
        _set(ws, f"BP{row}", _cv(line.mileage_rate))
    _set(ws, f"BU{row}", line.pov_reason)
    if line.other:
        _set(ws, f"CH{row}", _cv(line.other))
    _set(ws, f"CU{row}", line.purpose)
    # CB (mileage $) and CN (row total) are formula cells — do not overwrite


def _write_accounts(ws, intake: TravelIntake) -> None:
    for row, account in zip(range(34, 39), intake.account_codes[:5]):
        _set(ws, f"G{row}", account.work_order)
        _set(ws, f"U{row}", account.group)
        _set(ws, f"AA{row}", account.org_code)
        _set(ws, f"AJ{row}", account.object_code)
        _set(ws, f"BH{row}", account.work_op)
        _set(ws, f"BN{row}", account.bal_sheet)
        if account.amount:
            _set(ws, f"BT{row}", _cv(account.amount))


def _write_other_expenses(ws, items: List[OtherExpense]) -> None:
    for row, item in zip(range(34, 40), items[:6]):
        _set(ws, f"BZ{row}", item.date)
        _set(ws, f"CF{row}", item.paid_to)
        _set(ws, f"CY{row}", item.for_what)
        if item.amount:
            _set(ws, f"DS{row}", _cv(item.amount))


# ---------------------------------------------------------------------------
# Main function
# ---------------------------------------------------------------------------

def fill_expense_voucher_xlsx(
    intake: TravelIntake,
    template_path: str | Path,
    output_path: str | Path,
) -> Path:
    template_path = Path(template_path)
    output_path = Path(output_path)

    wb = openpyxl.load_workbook(str(template_path))
    ws = wb["133-103"] if "133-103" in wb.sheetnames else wb.active

    # Traveler header — page 1
    t = intake.traveler
    _set(ws, "A5",  t.name_last_first_initial or t.name)
    _set(ws, "AY5", t.employee_id)
    _set(ws, "CH5", t.official_station)
    _set(ws, "A7",  t.address)
    _set(ws, "AY7", t.city)
    _set(ws, "BP7", t.state)
    _set(ws, "BU7", t.zip_code)
    _set(ws, "CH7", t.official_residence)

    # Last travel date, regular hours, remarks
    if intake.daily_expenses:
        _set(ws, "A28", max(x.date for x in intake.daily_expenses))
    _set(ws, "Z28", t.regular_work_hours)
    _set(ws, "A30", intake.remarks)

    # Daily expense rows
    first_page_rows  = list(range(10, 22))   # rows 10–21  (12 rows)
    second_page_rows = list(range(48, 81))   # rows 48–80  (33 rows)

    page1_lines = intake.daily_expenses[: len(first_page_rows)]
    page2_lines = intake.daily_expenses[len(first_page_rows) : len(first_page_rows) + len(second_page_rows)]

    for row, line in zip(first_page_rows, page1_lines):
        _write_daily_line(ws, row, line)
    for row, line in zip(second_page_rows, page2_lines):
        _write_daily_line(ws, row, line)

    # Travel advance goes into BV27; the template formula CN27=CN25-BV27 computes the net total
    if intake.travel_advance:
        _set(ws, "BV27", _cv(intake.travel_advance))

    # Account codes and other expense receipts
    _write_accounts(ws, intake)
    _write_other_expenses(ws, intake.other_expenses)

    # Page 2 traveler header (repeat)
    _set(ws, "A45",  t.name_last_first_initial or t.name)
    _set(ws, "AY45", t.employee_id)
    if intake.daily_expenses:
        _set(ws, "BU45", max(x.date for x in intake.daily_expenses))

    # Signature date
    _set(ws, "AI43", intake.signature_date)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
